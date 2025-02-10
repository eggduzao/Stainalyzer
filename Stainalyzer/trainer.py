
"""
Trainer
-------

Placeholder.

"""

############################################################################################################
### Import
############################################################################################################

import os
import sys
import cv2
import random
import openpyxl
import numpy as np
from scipy.linalg import sqrtm
from collections import OrderedDict

from .utils import ColorConverter, PlottingUtils, TripleInterval
from .preprocessor import ImagePreprocessor
from .distribution import GaussianDistribution
from .filters import DABFilters, Mask

############################################################################################################
### Constants
############################################################################################################

# Constants
SEED = 1987
random.seed(SEED)
np.random.seed(SEED)

############################################################################################################
### Classes
############################################################################################################

class DABDistribution:
    """
    A class to represent a DAB (3,3'-diaminobenzidine) distribution, utilizing an instance of the GaussianDistribution class.

    Attributes:
    -----------
    gaussian_distribution : GaussianDistribution
        An instance of the GaussianDistribution class to model the DAB distribution.

    Methods:
    --------
    get_mean():
        Retrieves the mean vector of the Gaussian distribution.
    
    set_mean(mean):
        Sets the mean vector of the Gaussian distribution.
    
    get_covariance():
        Retrieves the covariance matrix of the Gaussian distribution.
    
    set_covariance(covariance):
        Sets the covariance matrix of the Gaussian distribution.

    get_mean_and_covariance():
        Returns both the mean vector and covariance matrix as a tuple.

    set_mean_and_covariance(mean, covariance):
        Sets both the mean vector and the covariance matrix of the Gaussian distribution.

    _wasserstein_distance(mean1, cov1, mean2, cov2):
        Computes the 2-Wasserstein distance between two Gaussian distributions.

    _kl_divergence(mean1, cov1, mean2, cov2):
        Computes the Kullback-Leibler divergence between two Gaussian distributions.

    _bhattacharyya_distance(mean1, cov1, mean2, cov2):
        Computes the Bhattacharyya distance between two Gaussian distributions.

    add(data):
        Adds new data points to the distribution and refits the GaussianDistribution instance.
    
    distance(other_gaussian_distribution):
        Calculates the distance to another GaussianDistribution instance.

    __repr__():
        Returns a detailed string representation of the DABDistribution object.

    __str__():
        Returns a user-friendly string representation of the DABDistribution object.
    """

    def __init__(self, gaussian_distribution=None, mean=None, covariance=None, n_samples=None):
        """
        Initializes a DABDistribution object with a GaussianDistribution instance, mean, and covariance.

        Parameters:
        -----------
        gaussian_distribution : GaussianDistribution, optional
            An existing GaussianDistribution instance. Defaults to None.
        mean : array-like, optional
            The mean vector for a new GaussianDistribution instance. Defaults to None.
        covariance : array-like, optional
            The covariance matrix for a new GaussianDistribution instance. Defaults to None.
        """
        if gaussian_distribution is not None:
            self.gaussian_distribution = gaussian_distribution
        elif mean is not None and covariance is not None and n_samples is not None:
            self.gaussian_distribution = GaussianDistribution(mean, covariance, n_samples)
        else:
            self.gaussian_distribution = None

    def get_mean(self):
        """
        Returns the mean vector of the Gaussian distribution.

        Returns:
        --------
        np.ndarray
            The mean vector.
        """
        return self.gaussian_distribution.mean if self.gaussian_distribution else None

    def set_mean(self, mean):
        """
        Sets the mean vector of the Gaussian distribution.

        Parameters:
        -----------
        mean : array-like
            The mean vector to set.
        """
        if self.gaussian_distribution:
            self.gaussian_distribution.mean = mean

    def get_covariance(self):
        """
        Returns the covariance matrix of the Gaussian distribution.

        Returns:
        --------
        np.ndarray
            The covariance matrix.
        """
        return self.gaussian_distribution.covariance if self.gaussian_distribution else None

    def set_covariance(self, covariance):
        """
        Sets the covariance matrix of the Gaussian distribution.

        Parameters:
        -----------
        covariance : array-like
            The covariance matrix to set.
        """
        if self.gaussian_distribution:
            self.gaussian_distribution.covariance = covariance

    def get_nsamples(self):
        """
        Returns the number of samples of the Gaussian distribution.

        Returns:
        --------
        int
            The number of samples.
        """
        return self.gaussian_distribution.n_samples if self.gaussian_distribution else 0

    def set_nsamples(self, n_samples):
        """
        Sets the number of samples of the Gaussian distribution.

        Parameters:
        -----------
        n_samples : array-like
            The number of samples to set.
        """
        if self.gaussian_distribution:
            self.gaussian_distribution.n_samples = n_samples

    def get_mean_covariance_and_nsamples(self):
        """
        Returns the mean vector, covariance matrix and number of samples of the Gaussian distribution.

        Returns:
        --------
        tuple
            A tuple containing the mean vector, the covariance matrix and the number of samples.
        """
        if self.gaussian_distribution:
            return self.gaussian_distribution.mean, self.gaussian_distribution.covariance, self.n_samples
        else:
            return None

    def set_mean_covariance_and_nsamples(self, mean, covariance, n_samples=0):
        """
        Sets the mean vector, the covariance matrix and number of samples of the Gaussian distribution.

        Parameters:
        -----------
        mean : array-like
            The mean vector to set.
        covariance : array-like
            The covariance matrix to set.
        n_samples : array-like
            The n_samples value to set.
        """
        if self.gaussian_distribution:
            self.gaussian_distribution = GaussianDistribution(mean, covariance, n_samples)

    def update(self, data):
        """
        Incrementally updates the Gaussian distribution using new data points without storing all previous data.
        
        Parameters
        ----------
        data : np.ndarray
            A 2D array where each row is a new data point.

        Notes
        -----
        This method updates both the mean vector and covariance matrix in an **incremental fashion** using Welford’s
        algorithm. This avoids numerical instability and is memory-efficient, making it suitable for large datasets.
        """

        data = np.array(data)
        n_new = data.shape[0]  # Number of new samples

        if self.gaussian_distribution is None:

            # Initialize with first batch of data
            mean = np.mean(data, axis=0)
            covariance = np.cov(data, rowvar=False)
            n_samples = np.int32(n_new)

            # Initialize Gaussian Distribution
            self.gaussian_distribution = GaussianDistribution(mean, covariance, n_samples)

        else:

            # Compute new mean incrementally
            new_mean = (
                (self.gaussian_distribution.n_samples * self.gaussian_distribution.mean + 
                 n_new * np.mean(data, axis=0)) / 
                (self.gaussian_distribution.n_samples + n_new)
            )

            # Compute covariance incrementally
            diff_old = self.gaussian_distribution.mean - new_mean
            diff_new = data - new_mean

            # Compute new covariance using Welford’s algorithm
            updated_cov = (
                (self.gaussian_distribution.n_samples * self.gaussian_distribution.covariance + 
                np.dot(diff_new.T, diff_new) + self.gaussian_distribution.n_samples * np.outer(diff_old, diff_old) /
                (self.gaussian_distribution.n_samples + n_new)) / (self.gaussian_distribution.n_samples + n_new - 1)
            )

            # Update stored values
            self.gaussian_distribution.mean = new_mean
            self.gaussian_distribution.covariance = updated_cov
            self.gaussian_distribution.n_samples += n_new

    def distance(self, other_gaussian_distribution, metric="Wasserstein"):
        """
        Calculates the distance to another GaussianDistribution instance.

        Parameters:
        -----------
        other_gaussian_distribution : GaussianDistribution
            Another GaussianDistribution instance to calculate the distance to.
        metric : str            
            Distance metric to use. Options:
                - "Wasserstein": 2-Wasserstein distance between distributions.
                - "KL-divergence": Kullback-Leibler divergence.
                - "Bhattacharyya": Bhattacharyya distance.
        Returns:
        --------
        float
            The calculated distance.
        """
        
        if self.gaussian_distribution is None or other_gaussian_distribution is None:
            raise ValueError("Both distributions must be initialized before computing distance.")

        mean1 = self.gaussian_distribution.mean
        cov1 = self.gaussian_distribution.covariance
        mean2 = other_gaussian_distribution.mean
        cov2 = other_gaussian_distribution.covariance

        if metric == "Wasserstein":
            return self._wasserstein_distance(mean1, cov1, mean2, cov2)

        elif metric == "KL-divergence":
            return self._kl_divergence(mean1, cov1, mean2, cov2)

        elif metric == "Bhattacharyya":
            return self._bhattacharyya_distance(mean1, cov1, mean2, cov2)

        else:
            raise ValueError("Invalid metric. Choose 'Wasserstein', 'KL-divergence', or 'Bhattacharyya'.")

    def _wasserstein_distance(self, mean1, cov1, mean2, cov2):
        """
        Computes the 2-Wasserstein distance between two Gaussian distributions.

        Returns:
        --------
        float
            Wasserstein distance.
        """
        mean_diff = np.linalg.norm(mean1 - mean2)
        cov_root = sqrtm(cov1 @ cov2)
        if np.iscomplexobj(cov_root):
            cov_root = cov_root.real  
        cov_diff = np.trace(cov1 + cov2 - 2 * cov_root)
        return np.sqrt(mean_diff**2 + cov_diff)

    def _kl_divergence(self, mean1, cov1, mean2, cov2):
        """
        Computes the Kullback-Leibler divergence between two Gaussian distributions.

        Returns:
        --------
        float
            KL divergence.
        """
        cov2_inv = np.linalg.inv(cov2)
        mean_diff = mean2 - mean1
        term1 = np.trace(cov2_inv @ cov1)
        term2 = mean_diff.T @ cov2_inv @ mean_diff
        term3 = np.log(np.linalg.det(cov2) / np.linalg.det(cov1))
        return 0.5 * (term1 + term2 - len(mean1) + term3)

    def _bhattacharyya_distance(self, mean1, cov1, mean2, cov2):
        """
        Computes the Bhattacharyya distance between two Gaussian distributions.

        Returns:
        --------
        float
            Bhattacharyya distance.
        """
        mean_diff = mean1 - mean2
        cov_avg = 0.5 * (cov1 + cov2)
        cov_avg_inv = np.linalg.inv(cov_avg)
        term1 = 0.125 * mean_diff.T @ cov_avg_inv @ mean_diff
        term2 = 0.5 * np.log(np.linalg.det(cov_avg) / np.sqrt(np.linalg.det(cov1) * np.linalg.det(cov2)))
        return term1 + term2

    def __repr__(self):
        """
        Returns a detailed string representation of the GaussianDistribution object.

        Returns:
        --------
        str
            A string representation of the object.
        """
        return f"DABDistribution(mean={self.get_mean()}, covariance={self.get_covariance()})"

    def __str__(self):
        """
        Returns a user-friendly string representation of the GaussianDistribution object.

        Returns:
        --------
        str
            A string describing the Gaussian distribution.
        """
        return (
            f"DAB Gaussian Distribution\n"
            f"Mean: {self.mean}\n"
            f"Covariance Matrix: {self.covariance}"
        )

class Trainer:
    """
    A class to handle the training process for a model using DABDistribution and a path to training images.

    Attributes:
    -----------
    dab_distribution : DABDistribution
        An instance of the DABDistribution class.
    training_image_path : str
        Path to the directory containing training images.
    severity : np.float32
        A float parameter (0 to 1) to control clipping points in the HSV image.

    Methods:
    --------
    get_dab_distribution():
        Retrieves the current DABDistribution instance.

    set_dab_distribution(dab_distribution):
        Sets the DABDistribution instance.

    get_training_image_path():
        Retrieves the training image path.

    set_training_image_path(training_image_path):
        Sets the training image path.

    get_severity():
        Retrieves the severity parameter.

    set_severity(severity):
        Sets the severity parameter.

    save_trained_model(file_path):
        Save this class to a file_path.

    load_trained_model(file_path):
        Loads this class from a file_path

    training_iteration():
        Placeholder for the training pipeline for a single image.

    train():
        Iterates over images in the training_image_path, calling training_iteration().
    """

    def __init__(self, dab_distribution=None, training_image_path=None, severity=None, root_name=None):
        """
        Initializes the Trainer object with a DABDistribution instance, training image path, and severity parameter.

        Parameters:
        -----------
        dab_distribution : DABDistribution, optional
            An instance of the DABDistribution class. Defaults to None.
        training_image_path : str, optional
            Path to the directory containing training images. Defaults to None.
        severity : np.float32, optional
            A float parameter (0 to 1) to control clipping points. Defaults to None.
        """

        self.dab_distribution = dab_distribution if dab_distribution is not None else DABDistribution()
        self.training_image_path = training_image_path if training_image_path is not None else None
        self.severity = severity if severity is not None else None
        self.root_name = root_name if root_name is not None else None
        self._plotting_utils = PlottingUtils()

    def get_dab_distribution(self):
        """
        Retrieves the current DABDistribution instance.

        Returns:
        --------
        DABDistribution
            The current DABDistribution instance.
        """
        return self.dab_distribution if self.dab_distribution else None

    def set_dab_distribution(self, dab_distribution):
        """
        Sets the DABDistribution instance.

        Parameters:
        -----------
        dab_distribution : DABDistribution
            The DABDistribution instance to set.
        """
        self.dab_distribution = dab_distribution

    def get_training_image_path(self):
        """
        Retrieves the training image path.

        Returns:
        --------
        str
            The path to the training images.
        """
        return self.training_image_path if self.training_image_path else None

    def set_training_image_path(self, training_image_path):
        """
        Sets the training image path.

        Parameters:
        -----------
        training_image_path : str
            The path to the directory containing training images.
        """
        self.training_image_path = training_image_path

    def get_severity(self):
        """
        Retrieves the severity parameter.

        Returns:
        --------
        np.float32
            The severity parameter.
        """
        return self.severity if self.severity else None

    def set_severity(self, severity):
        """
        Sets the severity parameter.

        Parameters:
        -----------
        severity : np.float32
            The float parameter (0 to 1) to control clipping points.
        """
        self.severity = severity

    def save_trained_model(self, file_path):
        """
        Saves the trained model to a tab-separated text file, including the DAB distribution parameters 
        (mean and covariance), training image path, and severity.

        Parameters
        ----------
        file_path : str
            Path where the model should be saved.

        Raises
        ------
        ValueError
            If the DAB distribution is not initialized before saving.
        """
        if self.dab_distribution is None or self.dab_distribution.gaussian_distribution is None:
            raise ValueError("DAB distribution must be initialized before saving the model.")

        mean = self.dab_distribution.get_mean()
        covariance = self.dab_distribution.get_covariance()
        nsamples = self.dab_distribution.get_nsamples()

        # Ensure values are stored in a readable format
        mean_str = "\t".join(map(str, mean))
        covariance_str = "\n".join(["\t".join(map(str, row)) for row in covariance])
        nsamples_str = str(nsamples)

        with open(file_path, "w") as f:
            f.write("# DAB Distribution Model\n")
            f.write(f"Training Path:\t{self.training_image_path}\n")
            f.write(f"Severity:\t{self.severity}\n")
            f.write(f"Mean:\t{mean_str}\n")
            f.write(f"Nsamples:\t{nsamples_str}\n")
            f.write("Covariance:\n")
            f.write(f"{covariance_str}\n")

    def load_trained_model(self, file_path):
        """
        Loads a trained model from a saved file, reconstructing the Trainer instance.

        Parameters
        ----------
        file_path : str
            Path to the saved model file.

        Raises
        ------
        FileNotFoundError
            If the specified file does not exist.
        ValueError
            If the file format is incorrect or missing required parameters.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        training_image_path = None
        severity = None
        mean = None
        covariance = []

        with open(file_path, "r") as f:
            lines = f.readlines()

        for line in lines:
            line = line.strip()

            if line.startswith("Training Path:"):
                training_image_path = line.split("\t")[1]

            elif line.startswith("Severity:"):
                severity = np.float32(line.split("\t")[1])

            elif line.startswith("Mean:"):
                mean = np.array([float(x) for x in line.split("\t")[1:]])

            elif line.startswith("Nsamples:"):
                nsamples = np.int32(line.split("\t")[1])

            elif line.startswith("Covariance:"):
                # All following lines are covariance matrix rows
                covariance_start_index = lines.index(line) + 1
                for cov_line in lines[covariance_start_index:]:
                    covariance.append([float(x) for x in cov_line.split("\t")])

        # Convert to numpy array
        if mean is None or len(covariance) == 0:
            raise ValueError("Invalid model file: Missing required parameters.")

        covariance = np.array(covariance)
        
        # Create a new DABDistribution instance
        dab_distribution = DABDistribution(mean=mean, covariance=covariance, n_samples=nsamples)

        # Reconstruct and return the Trainer instance
        self.dab_distribution = dab_distribution
        self.training_image_path = training_image_path
        self.severity = severity

    def compute_image_statistics(image):
        """
        Computes the mean color and covariance matrix of an image.

        Parameters
        ----------
        image : numpy.ndarray
            An image in OpenCV format (H, W, C) where C = 3 (BGR).

        Returns
        -------
        mean_vector : numpy.ndarray
            The mean value of each color channel (B, G, R).
        covariance_matrix : numpy.ndarray
            The 3x3 covariance matrix of color channels.
        """

        # Reshape image into a 2D array (N_pixels, 3) where each row is (B, G, R)
        pixels = image.reshape(-1, 3).astype(np.float32)  # Convert to float for accuracy

        # Compute the mean vector (average color per channel)
        mean_vector = np.mean(pixels, axis=0)

        # Compute the covariance matrix (color channel relationships)
        covariance_matrix = np.cov(pixels, rowvar=False)

        return mean_vector, covariance_matrix

    def get_masks(self, mask_level, hsv_image):
        """
        This method will process a single training iteration using the current image and distribution.

        Parameters
        ----------
        mask_level : int
            The mas ridigity, 1 for high to 3 for low.

        Returns:
        --------
        foreground_mask : np.float32
            The severity parameter.
        tissue_mask : np.float32
            The severity parameter.
        background_mask: np.float32
            The severity parameter.

        Raises
        ------
        ValueError
            If the specified mask level is not within 1 and 5.
        """

        if mask_level not in range(1, 4):
            raise ValueError(f"The mask level {mask_level} does not exist.")

        # Foreground Masks
        foreground_mask_1 = ( # Level 1 Foreground Mask
            cv2.bitwise_or(
                cv2.inRange(hsv_image, np.array([0, 0, 0]), np.array([30, 255, 160])),  # DAB brown, dark red, dark yellow
                cv2.inRange(hsv_image, np.array([175, 30, 50]), np.array([180, 255, 140]))  # Very dark plum-red
            )
        )
        foreground_mask_2 = ( # Level 2 Foreground Mask
            cv2.bitwise_or(
                cv2.inRange(hsv_image, np.array([0, 0, 0]), np.array([40, 255, 200])),  # DAB brown, dark red, yellow
                cv2.inRange(hsv_image, np.array([170, 0, 0]), np.array([180, 255, 200]))  # Slightly lighted plum-red-pink
            )
        )
        foreground_mask_3 = ( # Level 3 Foreground Masks
            cv2.bitwise_or(
                cv2.inRange(hsv_image, np.array([0, 0, 0]), np.array([50, 255, 255])),  # DAB brown, Dark Red, Yellow
                cv2.inRange(hsv_image, np.array([160, 0, 0]), np.array([180, 255, 255]))  # Plum, Red & Pink
            )
        )

        # Tissue Masks
        tissue_mask_1 = ( # Level 1 Tissue Mask
            cv2.bitwise_or(
                cv2.inRange(hsv_image, np.array([80, 0, 0]), np.array([100, 100, 100])),  # Darker Silver-Blue
                cv2.inRange(hsv_image, np.array([100, 0, 0]), np.array([120, 100, 100]))  # Darker Blue
            )
        )
        tissue_mask_2 = ( # Level 2 Tissue Mask
            cv2.bitwise_or(
                cv2.inRange(hsv_image, np.array([60, 0, 0]), np.array([100, 100, 100])),  # Darker Silver-Blue-Green
                cv2.inRange(hsv_image, np.array([100, 0, 0]), np.array([140, 100, 100]))  # Darker Blue
            )
        )
        tissue_mask_3 = ( # Level 3 Tissue Mask
            cv2.bitwise_or(
                cv2.inRange(hsv_image, np.array([50, 0, 0]), np.array([100, 100, 100])),  # Darker Silver-Blue
                cv2.inRange(hsv_image, np.array([100, 0, 0]), np.array([160, 100, 100]))  # Darker Blue
            )
        )

        # Background Masks
        background_mask_1 = ( # Level 1 Background Mask
            cv2.bitwise_or(
                cv2.inRange(hsv_image, np.array([80, 100, 100]), np.array([100, 255, 255])),  # Light Silver-Blue
                cv2.inRange(hsv_image, np.array([100, 100, 100]), np.array([120, 255, 255]))  # Light Cyan
            )
        )
        background_mask_2 = ( # Level 2 Background Mask
            cv2.bitwise_or(
                cv2.inRange(hsv_image, np.array([60, 100, 100]), np.array([100, 255, 255])),  # Silver-Blue-Green
                cv2.inRange(hsv_image, np.array([100, 100, 100]), np.array([140, 255, 255]))  # Cyan-Blue slide
            )
        )
        background_mask_3 = ( # Level 3 Background Mask
            cv2.bitwise_or(
                cv2.inRange(hsv_image, np.array([50, 100, 100]), np.array([100, 255, 255])),  # Dark& Light Silver-Blue-Green
                cv2.inRange(hsv_image, np.array([100, 100, 100]), np.array([160, 255, 255]))  # Cyan-Blue darker slide
            )
        )

        # Mask assignments
        if mask_level == 1:
            foreground_mask = foreground_mask_1
            tissue_mask = tissue_mask_1
            background_mask = background_mask_1
        elif mask_level == 2:
            foreground_mask = foreground_mask_2
            tissue_mask = tissue_mask_2
            background_mask = background_mask_2
        elif mask_level == 3:
            foreground_mask = foreground_mask_3
            tissue_mask = tissue_mask_3
            background_mask = background_mask_3
        else:
            raise ValueError(f"The mask level {mask_level} does not exist.")

        return foreground_mask, tissue_mask, background_mask

    def remove_annotation(self, image, threshold=3, color_space="BGR"):
        """
        This method removes annotation (pure black) from the image by substituting it by its
        nearest neighbor color.

        Parameters
        ----------
        name : str
            The current mask's name.
        color_space : str
            The color mode of the image. (Defaults to 'BGR')

        Returns
        ------
        weighted_vec : Tuple
            A tuple with three weights that sum to one.
        """
        image_preprocessor = ImagePreprocessor(image)
        image_preprocessor.replace_black_pixels(black_threshold=threshold, color_space=color_space)
        return image_preprocessor.original_image

    def get_mask_weights(self, name):
        """
        This method returns the corresponding tuple with three weights that sum to one, in the format:
        (foreground_weight, tissue_weight, background_weight)

        Parameters
        ----------
        name : str
            The current mask's name.

        Returns
        ------
        weighted_vec : Tuple
            A tuple with three weights that sum to one.
        """

        weighted_vec = (0.0, 0.0, 0.0)

        if name == "00_WHITE":
            weighted_vec =  (0.0, 0.0, 1.0)
        elif name == "01_BLACK":
            weighted_vec =  (0.0, 0.0, 1.0)
        elif name == "02_GRAY":
            weighted_vec =  (0.0, 0.5, 0.5)
        else:
            for i in range(0, 180, 10):
                curr_name_hh = f"HH{int(i/10)}"
                curr_name_hl = f"HL{int(i/10)}"
                curr_name_lh = f"LH{int(i/10)}"
                curr_name_ll = f"LL{int(i/10)}"
                if name == curr_name_hh:
                    if i in [0, 10, 20, 170]:
                        weighted_vec =  (1.0, 0.0, 0.0)
                    elif i == 30:
                        weighted_vec =  (0.6, 0.4, 0.0)
                    elif i in [40, 140, 150]:
                        weighted_vec =  (0.3, 0.7, 0.0)
                    elif i in [50, 120, 130]:
                        weighted_vec =  (0.0, 1.0, 0.0)
                    elif i in [60, 70, 100, 110]:
                        weighted_vec =  (0.0, 0.6, 0.4)
                    elif i in [80, 90]:
                        weighted_vec =  (0.0, 0.5, 0.5)
                    elif i == 160:
                        weighted_vec =  (0.5, 0.5, 0.0)
                elif name == curr_name_hl:
                    if i in [0, 10, 20, 170]:
                        weighted_vec =  (1.0, 0.0, 0.0)
                    elif i == 30:
                        weighted_vec =  (0.6, 0.4, 0.0)
                    elif i == 40:
                        weighted_vec =  (0.3, 0.7, 0.0)
                    elif i in [50, 130]:
                        weighted_vec =  (0.0, 1.0, 0.0)
                    elif i in [60, 70, 110, 120]:
                        weighted_vec =  (0.0, 0.6, 0.4)
                    elif i in [80, 90, 100]:
                        weighted_vec =  (0.0, 0.5, 0.5)
                    elif i in [140, 150]:
                        weighted_vec =  (0.2, 0.8, 0.0)
                    elif i == 160:
                        weighted_vec =  (0.5, 0.5, 0.0)
                elif name == curr_name_lh:
                    if i in [0, 10, 20, 170]:
                        weighted_vec =  (1.0, 0.0, 0.0)
                    elif i == 30:
                        weighted_vec =  (0.9, 0.1, 0.0)
                    elif i == 40:
                        weighted_vec =  (0.3, 0.4, 0.3)
                    elif i == 50:
                        weighted_vec =  (0.2, 0.4, 0.4)
                    elif i in [60, 120]:
                        weighted_vec =  (0.0, 0.6, 0.4)
                    elif i in [70, 80, 90, 100, 110]:
                        weighted_vec =  (0.0, 0.5, 0.5)
                    elif i == 130:
                        weighted_vec =  (0.0, 1.0, 0.0)
                    elif i == 140:
                        weighted_vec =  (0.2, 0.8, 0.0)
                    elif i == 150:
                        weighted_vec =  (0.5, 0.5, 0.0)
                    elif i == 160:
                        weighted_vec =  (0.8, 0.2, 0.0)
                elif name == curr_name_ll:
                    if i in [0, 10, 20, 170]:
                        weighted_vec =  (1.0, 0.0, 0.0)
                    elif i == 30:
                        weighted_vec =  (0.9, 0.1, 0.0)
                    elif i in [40, 160]:
                        weighted_vec =  (0.5, 0.5, 0.0)
                    elif i == 50:
                        weighted_vec =  (0.2, 0.6, 0.2)
                    elif i in [60, 70]:
                        weighted_vec =  (0.0, 1.0, 0.0)
                    elif i == 80:
                        weighted_vec =  (0.0, 0.6, 0.4)
                    elif i in [90, 100]:
                        weighted_vec =  (0.0, 0.5, 0.5)
                    elif i in [110, 120, 130]:
                        weighted_vec =  (0.0, 1.0, 0.0)
                    elif i == 140:
                        weighted_vec =  (0.2, 0.8, 0.0)
                    elif i == 150:
                        weighted_vec =  (0.5, 0.5, 0.0)
                    elif i == 160:
                        weighted_vec =  (0.8, 0.2, 0.0)

        return weighted_vec

    def training_iteration(self,
                           original_image,
                           iteration, ksize=(9, 9),
                           sigmaX=2,
                           image_mode="BGR",
                           plot_location=False,
                           file_name=None,
                           workbook=None,
                           white_threshold=2,
                           black_threshold=7,
                           gray_threshold=10):
        """
        This method will process a single training iteration using the current image and distribution.

        Parameters
        ----------
        original_image : np.ndarray
            The input image, typically read using OpenCV.
        iteration : int
            The iteration number.
        ksize : tuple
            Gausian Blur kernel size parameter. (default is (9, 9))
        sigmaX : int
            Gaussian Blur sigma parameter. (default is 2)
        mode : str, optional
            The color space mode of the original_image (default is "BGR").
        plot_location : str
            Path to plot preprocessed and masked images side by side with original images

        Raises
        ------
        ValueError
            If the specified file is not properly formatted.
        """

        if not isinstance(original_image, np.ndarray) or original_image.ndim != 3 or original_image.shape[2] != 3:
            raise ValueError(f"Training image '{original_image}' is not properly formatted.")

        ########################################################################################################
        # Step 1: Convert image to BGR and remove annotation
        ########################################################################################################

        original_image = ColorConverter.convert_any_to_bgr(original_image, image_mode=image_mode)
        original_image = self.remove_annotation(original_image, threshold=3, color_space="BGR")

        ########################################################################################################
        # Step 2: Convert image to HSV
        ########################################################################################################

        hsv_image = cv2.cvtColor(original_image, cv2.COLOR_BGR2HSV)

        ########################################################################################################
        # Step 3: Apply Gamma+Alpha Filters
        ########################################################################################################

        # Create Filter Class
        dfh = DABFilters()

        # Estimate & apply gamma filters
        hsv_gamma_value = dfh.gamma_histogram_analysis(hsv_image, color_space="HSV")
        hsv_gamma_image = dfh.gamma_correction(hsv_image, hsv_gamma_value, color_space="HSV")

        # Estimate & apply alpha filters
        hsv_alpha_value = dfh.alpha_histogram_analysis(hsv_gamma_image, color_space="HSV")
        hsv_alpha_image = dfh.alpha_correction(hsv_gamma_image, hsv_gamma_value, color_space="HSV")

        ########################################################################################################
        # Step 4: Creating Masks
        ########################################################################################################

        # Mask ordered dictionary
        mask_dictionary = OrderedDict()

        # White, Black and Gray HSV intervals
        white_interval = TripleInterval((0, 0, 255-white_threshold), (180, white_threshold, 255))
        black_interval = TripleInterval((0, 0, 0), (180, 255, black_threshold))
        gray_interval = TripleInterval((0, 0, 0), (180, gray_threshold, 255))

        # White, Black and Gray masks
        white_mask = Mask("00_WHITE", self.get_mask_weights("00_WHITE"), mask=white_interval, image=hsv_alpha_image)
        black_mask = Mask("01_BLACK", self.get_mask_weights("01_BLACK"), mask=black_interval, image=hsv_alpha_image)
        gray_mask = Mask("02_GRAY", self.get_mask_weights("02_GRAY"), mask=gray_interval, image=hsv_alpha_image)

        # Put in dictionary
        mask_dictionary[white_mask.name] = white_mask
        mask_dictionary[black_mask.name] = black_mask
        mask_dictionary[gray_mask.name] = gray_mask

        # Iteration for further dictionaries
        background_pixels = (0, 0, 255)
        for i in range(0, 180, 10):

            # Create names
            curr_name_ll = f"LL{int(i/10)}"
            curr_name_lh = f"LH{int(i/10)}"
            curr_name_hl = f"HL{int(i/10)}"
            curr_name_hh = f"HH{int(i/10)}"

            # Create intervals
            curr_interval_ll = TripleInterval((i, gray_threshold, black_threshold), (i+10, 128, 128))
            curr_interval_lh = TripleInterval((i, gray_threshold, 127), (i+10, 128, 255-white_threshold))
            curr_interval_hl = TripleInterval((i, 127, black_threshold), (i+10, 255, 128))
            curr_interval_hh = TripleInterval((i, 127, 127), (i+10, 255, 255-white_threshold))

            # Create masks
            curr_mask_ll = Mask(curr_name_ll, self.get_mask_weights(curr_name_ll),
                                mask=curr_interval_ll, image=hsv_alpha_image, background_pixels=background_pixels)
            curr_mask_lh = Mask(curr_name_lh, self.get_mask_weights(curr_name_lh),
                                mask=curr_interval_lh, image=hsv_alpha_image, background_pixels=background_pixels)
            curr_mask_hl = Mask(curr_name_hl, self.get_mask_weights(curr_name_hl),
                                mask=curr_interval_hl, image=hsv_alpha_image, background_pixels=background_pixels)
            curr_mask_hh = Mask(curr_name_hh, self.get_mask_weights(curr_name_hh),
                                mask=curr_interval_hh, image=hsv_alpha_image, background_pixels=background_pixels)

            # Put in dictionaries
            mask_dictionary[curr_mask_ll.name] = curr_mask_ll
            mask_dictionary[curr_mask_lh.name] = curr_mask_lh
            mask_dictionary[curr_mask_hl.name] = curr_mask_hl
            mask_dictionary[curr_mask_hh.name] = curr_mask_hh

        ########################################################################################################
        # Step 5: Calculate Results
        ########################################################################################################

        # Result images
        result_image_foreground = np.full_like(original_image, background_pixels, dtype=np.uint8)
        result_image_tissue = np.full_like(original_image, background_pixels, dtype=np.uint8)
        result_image_background = np.full_like(original_image, background_pixels, dtype=np.uint8)
        result_image_vec = [result_image_foreground, result_image_tissue, result_image_background]

        # Total weighted pixels for foreground, tissue and background
        total_weighted_pixel_vec = [0, 0, 0]
        total_pixel_vec = [0, 0, 0]

        # Process Each Mask
        for k, (name, mask) in enumerate(mask_dictionary.items()):

            # Calculate each score given the white mask
            weighted_score = mask.weighted_score()
            tpixel = sum(weighted_score)
            if(tpixel <= 0):
                continue
            total_weighted_pixel_vec = tuple(a + b for a, b in zip(total_weighted_pixel_vec, weighted_score))          
            total_pixel_vec = tuple(total_pixel_vec[e] + tpixel if weighted_score[e]>0 else total_pixel_vec[e] for e in range(0, 3))

            # Update image
            for img, wgt in zip(result_image_vec, mask.weight_tuple):
                mask.put(img, proportion=wgt)

        # Compute Mean & Covariance Matrices
        meanvec = [Mask.compute_mean(img) for img in result_image_vec]
        covvec = [Mask.compute_covariance(img) for img in result_image_vec]

        # Total Values
        total_weighted_pixel_vec_sum = sum(total_weighted_pixel_vec)
        total_pixel_vec_sum = sum(total_pixel_vec)

        # Write Summary Row in Excel
        summary_data = [

            # Total Forward Weighted Pixels, Total Tissue Weighted Pixels, Total Background Weighted Pixels
            *total_weighted_pixel_vec,

            # Weighted Forward to Tissue (%)
            round(
                (total_weighted_pixel_vec[0] / total_weighted_pixel_vec[1]) * 100, 4
            ) if total_weighted_pixel_vec[1] else 0,
            # Weighted Forward To Background (%)
            round(
                (total_weighted_pixel_vec[0] / total_weighted_pixel_vec[2]) * 100, 4
            ) if total_weighted_pixel_vec[2] else 0,
            # Weighted Forward To Tissue+Background (%)
            round(
                (total_weighted_pixel_vec[0] / (total_weighted_pixel_vec[1] + total_weighted_pixel_vec[2])) * 100, 4
            ) if total_weighted_pixel_vec[1] + total_weighted_pixel_vec[2] else 0,
            # Weighted Forward To All (%)
            round(
                (total_weighted_pixel_vec[0] / total_weighted_pixel_vec_sum) * 100, 4
            ) if total_weighted_pixel_vec_sum else 0,
            
            # Weighted Total (Pixels)
            total_weighted_pixel_vec_sum,

            # Total Forward (Pixels), Total Tissue (Pixels), Total Background (Pixels)
            *total_pixel_vec,

            # Forward to Tissue (%)
            round((total_pixel_vec[0] / total_pixel_vec[1]) * 100, 4) if total_pixel_vec[1] else 0,
            # Forward To Background (%)
            round((total_pixel_vec[0] / total_pixel_vec[2]) * 100, 4) if total_pixel_vec[2] else 0,
            # Forward To Tissue+Background (%)
            round(
                (total_pixel_vec[0] / (total_pixel_vec[1] + total_pixel_vec[2])) * 100, 4
            ) if total_pixel_vec[1] + total_pixel_vec[2] else 0,
            # Forward To All (%)
            round((total_pixel_vec[0] / total_pixel_vec_sum) * 100, 4) if total_pixel_vec_sum else 0,

            # Total (Pixels)
            total_pixel_vec_sum,

            # Forward HSV Distribution
            *meanvec[0], *covvec[0].flatten(),

            # Tissue HSV Distribution
            *meanvec[1], *covvec[1].flatten(),

            # Background HSV Distribution
            *meanvec[2], *covvec[2].flatten()
        ]

        # Writing to excel
        if workbook is not None:
            sheet = workbook.active
            for col, value in enumerate(summary_data, start=1):
                sheet.cell(row=iteration, column=col, value=value)

        # Save & Print Results
        output_file_name = os.path.join(plot_location, f"result_{file_name}.tiff")
        self._plotting_utils.save_picture_labels(output_file_name, *result_image_vec, pixel_counts=total_weighted_pixel_vec, 
                                                 n_cols=3, color_space="HSV", plot_grids=True)

        return iteration + 1

    def train(self, output_location=None):
        """
        This method will process a path, updating self.dab_distribution with each image's DAB content.

        Raises
        ------
        FileNotFoundError
            If the specified path does not exist.
        ValueError
            If the output location is incorrect and cannot be created.
        """

        if not os.path.exists(self.training_image_path):
            raise FileNotFoundError(f"Training image path '{self.training_image_path}' does not exist.")

        workbook = None
        excel_file_name = None
        if output_location is not None:

            if not os.path.exists(output_location):
                try:
                    os.makedirs(output_location, exist_ok=True)
                except Exception:
                    raise ValueError("Output location {output_location} is incorrect and cannot be created.")

            # Open Excel spreadsheet
            excel_file_name = os.path.join(output_location, "results.xlsx")
            
            if os.path.exists(excel_file_name):
                workbook = openpyxl.load_workbook(excel_file_name)
            else:
                # Create a new Excel file
                workbook = openpyxl.Workbook()
                workbook.save(excel_file_name)
                workbook.close()
                workbook = openpyxl.load_workbook(excel_file_name)

            # Writing first line header
            sheet = workbook.active
            data_to_write = ["Total Forward Weighted (Pixels)", "Total Tissue Weighted (Pixels)", "Total Background Weighted (Pixels)",
                             "Weighted Forward to Tissue (%)", "Weighted Forward To Background (%)", 
                             "Weighted Forward To Tissue+Background (%)", "Weighted Forward To All (%)", "Weighted Total (Pixels)",
                             "Total Forward (Pixels)", "Total Tissue (Pixels)", "Total Background (Pixels)",
                             "Forward to Tissue (%)", "Forward To Background (%)", "Forward To Tissue+Background (%)", 
                             "Forward To All (%)", "Total (Pixels)",
                             "Average Forward Hue (H)", "Average Forward Saturation (S)", "Average Forward Value (V)", 
                             "Cov Forward HH", "Cov Forward HS", "Cov Forward HV",
                             "Cov Forward SH", "Cov Forward SS", "Cov Forward SV",
                             "Cov Forward VH", "Cov Forward VS", "Cov Forward VV", 
                             "Average Tissue Hue (H)", "Average Tissue Saturation (S)", "Average Tissue Value (V)", 
                             "Cov Tissue HH", "Cov Tissue HS", "Cov Tissue HV",
                             "Cov Tissue SH", "Cov Tissue SS", "Cov Tissue SV",
                             "Cov Tissue VH", "Cov Tissue VS", "Cov Tissue VV", 
                             "Average Background Hue (H)", "Average Background Saturation (S)", "Average Background Value (V)", 
                             "Cov Background HH", "Cov Background HS", "Cov Background HV",
                             "Cov Background SH", "Cov Background SS", "Cov Background SV",
                             "Cov Background VH", "Cov Background VS", "Cov Background VV"]
            for col, value in enumerate(data_to_write, start=1):
                sheet.cell(row=1, column=col, value=value)

        counter = 2
        for dirpath, dirnames, filenames in os.walk(self.training_image_path):

            for image_file in filenames:

                # If not an image in TIFF format continue to next
                if not image_file.lower().endswith(".tiff"):
                    continue

                # Count Iteration
                counter += 1

                # Image path
                image_path = os.path.join(dirpath, image_file)
                after_root = image_path.split(self.root_name + os.sep, 1)[1]

                # Split the remaining path into components
                path_components = after_root.split(os.sep)
                image_file_name = os.path.splitext(image_file)[0]

                if excel_file_name is not None:
                    sheet = workbook.active
                    for col, path_component in enumerate(["File Name:", self.root_name] + path_components, start=1):
                        sheet.cell(row=counter, column=col, value=path_component)

                # Create output location tree-wise
                treewise_output_location = os.path.join(output_location, *path_components[:-1])
                try:
                    os.makedirs(treewise_output_location, exist_ok=True)
                except Exception:
                    raise ValueError("Output location {treewise_output_location} is incorrect and cannot be created.")

                # Perform training
                if os.path.isfile(image_path):
                    image = cv2.imread(image_path)
                    counter += 1
                    counter = self.training_iteration(image,
                                                      counter,
                                                      plot_location=treewise_output_location,
                                                      file_name=image_file_name,
                                                      workbook=workbook
                                                      )

                if excel_file_name is not None:
                    sheet = workbook.active
                    sheet.cell(row=counter, column=1, value=" ")

        if excel_file_name is not None:
            workbook.save(excel_file_name)
            workbook.close()

        if output_location is not None and self.dab_distribution.gaussian_distribution is not None:
            output_file_name = os.path.join(output_location, "trained_model.txt")
            self.save_trained_model(output_file_name)

"""

        wt = 2; bt = 7; gv = 10

        white_mask = cv2.inRange(hsv_alpha_image, np.array([0, 0, 255-wt]), np.array([180, wt, 255]))
        black_mask = cv2.inRange(hsv_alpha_image, np.array([0, 0, 0]), np.array([180, 255, bt]))
        gray_mask = cv2.inRange(hsv_alpha_image, np.array([0, 0, 0]), np.array([180, gv, 255]))

        lowS_lowV_mask = OrderedDict({f"LL{int(e/10)}": cv2.inRange(hsv_alpha_image, np.array([e, gv, bt]), np.array([e+10, 128, 128])) for e in range(0, 180, 10)})
        lowS_highV_mask = OrderedDict({f"LH{int(e/10)}": cv2.inRange(hsv_alpha_image, np.array([e, gv, 127]), np.array([e+10, 128, 255-wt])) for e in range(0, 180, 10)})
        highS_lowV_mask = OrderedDict({f"HL{int(e/10)}": cv2.inRange(hsv_alpha_image, np.array([e, 127, bt]), np.array([e+10, 255, 128])) for e in range(0, 180, 10)})
        highS_highV_mask = OrderedDict({f"HH{int(e/10)}": cv2.inRange(hsv_alpha_image, np.array([e, 127, 127]), np.array([e+10, 255, 255-wt])) for e in range(0, 180, 10)})

        im01 = ColorConverter.apply_masked_background(images=[hsv_alpha_image], masks=[white_mask], color_modes=["HSV"], background_color="white")[0]
        im02 = ColorConverter.apply_masked_background(images=[hsv_alpha_image], masks=[black_mask], color_modes=["HSV"], background_color="white")[0]
        im03 = ColorConverter.apply_masked_background(images=[hsv_alpha_image], masks=[gray_mask], color_modes=["HSV"], background_color="white")[0]

        lowS_lowV_img = OrderedDict({f"LL{int(e/10)}": ColorConverter.apply_masked_background(images=[hsv_alpha_image], masks=[lowS_lowV_mask[f"LL{int(e/10)}"]], color_modes=["HSV"], background_color="white")[0] for e in range(0, 180, 10)})
        lowS_highV_img = OrderedDict({f"LH{int(e/10)}": ColorConverter.apply_masked_background(images=[hsv_alpha_image], masks=[lowS_highV_mask[f"LH{int(e/10)}"]], color_modes=["HSV"], background_color="white")[0] for e in range(0, 180, 10)})
        highS_lowV_img = OrderedDict({f"HL{int(e/10)}": ColorConverter.apply_masked_background(images=[hsv_alpha_image], masks=[highS_lowV_mask[f"HL{int(e/10)}"]], color_modes=["HSV"], background_color="white")[0] for e in range(0, 180, 10)})
        highS_highV_img = OrderedDict({f"HH{int(e/10)}": ColorConverter.apply_masked_background(images=[hsv_alpha_image], masks=[highS_highV_mask[f"HH{int(e/10)}"]], color_modes=["HSV"], background_color="white")[0] for e in range(0, 180, 10)})

        hsv_image_list = [im01, im02, im03] + list(lowS_lowV_img.values()) + list(lowS_highV_img.values()) + list(highS_lowV_img.values()) + list(highS_highV_img.values())
        hsv_image_name = ["00_WHITE", "01_BLACK", "02_GRAY"] + list(lowS_lowV_img.keys()) + list(lowS_highV_img.keys()) + list(highS_lowV_img.keys()) + list(highS_highV_img.keys())

        # Counting total number of pixels
        background_color = (0, 0, 255)
        hsv_image_total_pixels = ColorConverter.count_non_masked_pixels(hsv_alpha_image, background_color)
        total_white_pixels = ColorConverter.count_color_pixels(hsv_alpha_image, background_color)
        hsv_image_total_pixels += total_white_pixels

        # Optional printing images
        if plot_location is not None and isinstance(plot_location, str):
            for (selected_image, selected_name) in zip(hsv_image_list, hsv_image_name):
                #if i == 0:
                #    lab = "hsv"
                #else:
                #    lab = "lab"
                output_file_name = os.path.join(plot_location, f"{iteration}_{selected_name}.tiff")
                if selected_name == "WHITE":
                    curr_pixels = total_white_pixels
                else:
                    curr_pixels = ColorConverter.count_non_masked_pixels(selected_image, background_color)
                pixel_counts = (curr_pixels, hsv_image_total_pixels)
                image_vec = (selected_image, hsv_alpha_image)
                self._plotting_utils.save_picture_labels(output_file_name, *image_vec, pixel_counts=pixel_counts,
                                                         n_cols=2, color_space="HSV", plot_grids=True)

        # Step 8: Extract H, S, V channels from hsv_selected and L, A, B from lab_selected
        #H, S, V = cv2.split(hsv_selected)  # Extract HSV channels
        #L, A, B = cv2.split(lab_selected)  # Extract LAB channels

        # Step 9: Normalize all channels to be between 0 and 1 (assuming normalize() exists)
        #H_norm = ColorConverter.normalize(H)
        #S_norm = ColorConverter.normalize(S)
        #V_norm = ColorConverter.normalize(V)
        #L_norm = ColorConverter.normalize(L)
        #A_norm = ColorConverter.normalize(A)
        #B_norm = ColorConverter.normalize(B)

        # Create a 6D feature vector with [H, S, V, L, A, B] for DABDistribution update
        #features_6D = np.column_stack((H_norm.flatten(), S_norm.flatten(), V_norm.flatten(),
        #                               L_norm.flatten(), A_norm.flatten(), B_norm.flatten()))

        # Update the DABDistribution with the extracted feature data
        #self.dab_distribution.update(features_6D)


class Trainer:
    
    A class to manage the training process for various distribution models.

    The Trainer class is responsible for:
    - Loading training data.
    - Preprocessing the data (e.g., CLAHE normalization, annotation removal).
    - Configuring and fitting models like DABDistribution or MultivariateLABDistribution.
    - Tracking and logging training progress.
    - Saving trained models for later use.

    Attributes:
        training_data (list or numpy.ndarray): The dataset used for training.
        preprocessed_data (list or numpy.ndarray): Preprocessed dataset ready for training.
        model (object): The model being trained (e.g., DABDistribution, MultivariateLABDistribution).
        config (dict): Configuration parameters for the training process.
        logger (object, optional): Logger for tracking training progress and results.
    

    def __init__(self, model=None, training_data=None, config=None, logger=None):
        "" "
        Initialize the Trainer class.

        Parameters:
            model (object, optional): The model to be trained. Can be an instance of
                                      `DABDistribution` or `MultivariateLABDistribution`.
                                      Default is None.
            training_data (list or numpy.ndarray, optional): Training dataset. Default is None.
            config (dict, optional): Configuration parameters for training. Default is None.
            logger (object, optional): Logger for tracking progress. Default is None.

        Attributes:
            training_data (list or numpy.ndarray): Dataset loaded into the trainer.
            preprocessed_data (list or numpy.ndarray): Preprocessed data for training.
            model (object): The model to be configured and trained.
            config (dict): Training configuration, including model and preprocessing parameters.
            logger (object): Logger instance for progress tracking and debugging.
        "" "
        # Initialize attributes
        self.training_data = training_data if training_data is not None else []
        self.preprocessed_data = None  # Placeholder for preprocessed data
        self.model = model  # Assign the provided model or None
        self.config = config if config is not None else {}  # Default to empty configuration
        self.logger = logger  # Assign the logger or None

        print(type(selected_image))
        print(selected_image.dtype)
        print(selected_image.shape)
        print(selected_image.ndim)
        print(selected_image.size)
        print(np.min(selected_image))
        print(np.max(selected_image)) 
        print(type(hsv_gamma_image))
        print(hsv_gamma_image.dtype)
        print(hsv_gamma_image.shape)
        print(hsv_gamma_image.ndim)
        print(hsv_gamma_image.size)
        print(np.min(hsv_gamma_image))
        print(np.max(hsv_gamma_image)) 

        # Ensure preprocessing configuration exists in config
        if "preprocessing" not in self.config:
            self.config["preprocessing"] = {
                "annotation_removal": True,
                "clahe": {"clip_limit": 2.0, "tile_grid_size": (8, 8)},
                "slic": {"n_segments": 250, "compactness": 10.0},
                "kmeans": {"n_clusters": 10},
            }

        # TODO - Validate input training data structure
        if self.training_data and not isinstance(self.training_data, (list, np.ndarray)):
            raise ValueError("training_data must be a list or numpy.ndarray.")

        # Log initialization
        if self.logger:
            self.logger.info(f"Trainer initialized with model: {self.model} and config: {self.config}")
        else:
            print(f"[TRAINER INIT]: Initialized with model: {self.model} and config: {self.config}")

    def load_data(self, data_path):
        "" "
        Load training data from a specified path.

        This method loads image files from the specified directory, applies necessary preprocessing,
        and extracts LAB color data to update the DABDistribution.

        Parameters:
            data_path (str): Path to the directory containing training images.

        Returns:
            None

        Raises:
            FileNotFoundError: If the specified path does not exist.
            ValueError: If no valid image files are found in the directory.
        "" "

        # Check if the path exists
        if not os.path.exists(data_path):
            raise FileNotFoundError(f"The specified data path does not exist: {data_path}")

        # Initialize a list to store LAB data from all training images
        all_lab_data = []

        # Iterate through all files in the directory
        for file_name in os.listdir(data_path):
            file_path = os.path.join(data_path, file_name)

            # Skip non-image files
            if not file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp')):
                continue

            # Load the image using OpenCV
            image = cv2.imread(file_path)

            if image is None:
                print(f"Warning: Could not load image file: {file_name}")
                continue

            # Apply preprocessing pipeline to the image
            # TODO - Ensure ImagePreprocessor is properly configured and available
            preprocessor = ImagePreprocessor()
            processed_image = preprocessor.remove_annotations(image)  # Remove annotations
            processed_image = preprocessor.apply_clahe(processed_image)  # Apply CLAHE

            # Convert the preprocessed image to LAB color space
            lab_image = cv2.cvtColor(processed_image, cv2.COLOR_BGR2LAB)

            # Flatten the LAB image to extract LAB pixels
            lab_data = lab_image.reshape(-1, 3)

            # Filter out background pixels (e.g., black background with LAB value (0, 128, 128))
            # TODO - Confirm the exact background filtering criteria
            non_background_mask = ~np.all(lab_data == [0, 128, 128], axis=1)
            lab_data = lab_data[non_background_mask]

            # Add the LAB data to the collection
            all_lab_data.append(lab_data)

        # Check if any valid LAB data was extracted
        if not all_lab_data:
            raise ValueError(f"No valid LAB data found in the specified directory: {data_path}")

        # Concatenate all LAB data into a single array
        combined_lab_data = np.vstack(all_lab_data)

        # Update the DAB distribution with the extracted LAB data
        # TODO - Ensure the DABDistribution instance is correctly initialized and accessible
        self.dab_distribution.add_data(combined_lab_data)

        print(f"Loaded and processed training data from: {data_path}")

    def preprocess_data(self, **kwargs):
        "" "
        Preprocess the training data to prepare it for the model.

        This method applies the preprocessing pipeline to the training data, including:
        - Annotation removal
        - CLAHE normalization

        Parameters:
            **kwargs: Additional parameters for preprocessing, such as:
                      - clahe_params: Dictionary of parameters for CLAHE.
                      - background_filter: Criteria for filtering out background pixels.

        Returns:
            None
        "" "

        if not hasattr(self, 'training_data') or self.training_data is None:
            raise ValueError("Training data is not loaded. Please load data using 'load_data' before preprocessing.")

        # Extract CLAHE parameters from kwargs if available
        clahe_params = kwargs.get('clahe_params', {})
        background_filter = kwargs.get('background_filter', [0, 128, 128])  # Default black background in LAB

        # Initialize a list to store preprocessed data
        preprocessed_data = []

        for image_data in self.training_data:
            # Assuming training_data is a list of images
            # TODO - Ensure `training_data` is structured correctly (e.g., list of numpy arrays)

            # Perform annotation removal
            # TODO - Implement or validate annotation removal in `ImagePreprocessor`
            image_preprocessor = ImagePreprocessor()
            processed_image = image_preprocessor.remove_annotations(image_data)

            # Apply CLAHE normalization
            # TODO - Implement or validate CLAHE in `ImagePreprocessor`
            processed_image = image_preprocessor.apply_clahe(processed_image, **clahe_params)

            # Convert processed image to LAB color space
            lab_image = cv2.cvtColor(processed_image, cv2.COLOR_BGR2LAB)

            # Filter out background pixels
            lab_data = lab_image.reshape(-1, 3)
            non_background_mask = ~np.all(lab_data == background_filter, axis=1)
            lab_data = lab_data[non_background_mask]

            # Add preprocessed data to the collection
            preprocessed_data.append(lab_data)

        # Store the preprocessed data for further steps
        self.preprocessed_data = np.vstack(preprocessed_data) if preprocessed_data else None

        if self.preprocessed_data is None or len(self.preprocessed_data) == 0:
            raise ValueError("No valid data found after preprocessing.")

        print("Preprocessing completed. Data is ready for training.")

    def configure_model(self, model_type, **kwargs):
        "" "
        Configure the model for training.

        This method initializes the specified model type and stores it in the trainer instance.
        It also supports additional configurations passed via keyword arguments.

        Parameters:
            model_type (str): The type of model to initialize. Options are:
                              - 'DABDistribution': Creates a DAB distribution model.
                              - 'MultivariateLABDistribution': Creates a multivariate LAB distribution model.
            **kwargs: Additional parameters for the model configuration (e.g., number of components for GMM).

        Raises:
            ValueError: If the specified model type is not supported.

        Returns:
            None
        "" "

        # Validate the model_type input
        supported_models = ['DABDistribution', 'MultivariateLABDistribution']
        if model_type not in supported_models:
            raise ValueError(f"Unsupported model_type: '{model_type}'. Supported models: {supported_models}")

        # Configure the specified model type
        if model_type == 'DABDistribution':
            # TODO - Validate or initialize any required settings for DABDistribution
            self.model = DABDistribution(**kwargs)

        elif model_type == 'MultivariateLABDistribution':
            # TODO - Validate or initialize any required settings for MultivariateLABDistribution
            self.model = MultivariateLABDistribution(**kwargs)

        # Store the model type for reference
        self.model_type = model_type

        print(f"Model configured successfully: {model_type}")

    def fit(self, **kwargs):
        "" "
        Train the model on the loaded and preprocessed data.

        This method fits the selected model (DABDistribution or MultivariateLABDistribution) using the
        preprocessed training data.

        Parameters:
            **kwargs: Additional parameters for the fitting process (e.g., number of components for GMM).

        Raises:
            ValueError: If no training data is loaded or preprocessed.
            ValueError: If the model is not configured before calling this method.

        Returns:
            None
        "" "

        # Ensure the data is loaded and preprocessed
        if not hasattr(self, 'preprocessed_data') or self.preprocessed_data is None:
            raise ValueError("No preprocessed data available. Please preprocess the data before fitting the model.")

        # Ensure the model is configured
        if not hasattr(self, 'model') or self.model is None:
            raise ValueError("Model is not configured. Please configure the model before fitting.")

        # Fit the model using the preprocessed data
        if isinstance(self.model, DABDistribution):
            # Fitting the DABDistribution with preprocessed LAB data
            # TODO - Ensure the data format matches the expected shape
            self.model.add_data(self.preprocessed_data, **kwargs)

        elif isinstance(self.model, MultivariateLABDistribution):
            # Fitting the MultivariateLABDistribution directly
            # TODO - Validate input data shape and other fitting prerequisites
            self.model.fit(self.preprocessed_data, **kwargs)

        # Log a success message
        print("Model fitting completed successfully.")

    def save_model(self, file_path):
        "" "
        Save the trained model to a file.

        This method serializes and saves the currently configured and trained model to a specified file path.

        Parameters:
            file_path (str): Path to save the model file.

        Raises:
            ValueError: If the model is not configured or not fitted.

        Returns:
            None
        "" "

        # Ensure the model exists and is fitted
        if not hasattr(self, 'model') or self.model is None:
            raise ValueError("No model is configured. Please configure and fit a model before saving.")
        if not hasattr(self.model, 'is_fitted') or not self.model.is_fitted:
            raise ValueError("The model is not fitted. Please fit the model before saving.")

        # TODO - Ensure file_path is valid and writable
        self.model.save(file_path)
        print(f"Model successfully saved to {file_path}.")

    def log_progress(self, message):
        "" "
        Log a message about the training progress.

        This method provides a consistent way to log progress messages during the training process.

        Parameters:
            message (str): The message to log.

        Returns:
            None
        "" "

        # TODO - Replace with a proper logging mechanism if needed (e.g., Python's logging module)
        print(f"[TRAINER LOG]: {message}")

    def __repr__(self):
        "" "
        Representation of the Trainer class.

        Returns:
            str: Summary of the Trainer's current state, including the model type and data information.
        "" "
        
        model_info = f"Model: {self.model.__class__.__name__}" if hasattr(self, 'model') and self.model else "No model configured"
        data_info = f"Data: {len(self.preprocessed_data)} samples" if hasattr(self, 'preprocessed_data') and self.preprocessed_data is not None else "No data loaded"
        return f"Trainer({model_info}, {data_info})"
"""
