from pathlib import Path
from openpyxl import load_workbook

# Definindo prefixo
prefix = Path("/Users/egg/Projects/Stainalyzer/data/results/")

# Lista de pastas
folder_list = [
    prefix / "Neila_DAB/DAB_Grupo_Controle/GRUPO_CONTROLE_PE_19_12_2024/",
    prefix / "Neila_DAB/DAB_Grupo_Controle/GRUPO_CONTROLE_PE_ACE2/",
    prefix / "Neila_DAB/DAB_Grupo_Controle/GRUPO_CONTROLE_PE_CD28/",
    prefix / "Neila_DAB/DAB_Grupo_Controle/GRUPO_CONTROLE_PE_CTLA4/",
    prefix / "Neila_DAB/DAB_Grupo_Controle/GRUPO_CONTROLE_PE_HLAG2/",
    prefix / "Neila_DAB/DAB_Grupo_Controle/GRUPO_CONTROLE_PE_HLAG5/",
    prefix / "Neila_DAB/DAB_Grupo_Controle/GRUPO_CONTROLE_PE_PD1/",
    prefix / "Neila_DAB/DAB_Grupo_Controle/GRUPO_CONTROLE_PE_PDL1/",
    prefix / "Neila_DAB/DAB_Grupo_Controle/GRUPO_CONTROLE_PE_SPIKE/",
    prefix / "Neila_DAB/DAB_IMIP_Tratamento/IMIP_ACE2/",
    prefix / "Neila_DAB/DAB_IMIP_Tratamento/IMIP_CD28/",
    prefix / "Neila_DAB/DAB_IMIP_Tratamento/IMIP_CTLA4/",
    prefix / "Neila_DAB/DAB_IMIP_Tratamento/IMIP_FOTOS_NOVAS/",
    prefix / "Neila_DAB/DAB_IMIP_Tratamento/IMIP_HLAG2/",
    prefix / "Neila_DAB/DAB_IMIP_Tratamento/IMIP_HLAG5/",
    prefix / "Neila_DAB/DAB_IMIP_Tratamento/IMIP_PD1/",
    prefix / "Neila_DAB/DAB_IMIP_Tratamento/IMIP_PDL1/",
    prefix / "Neila_DAB/DAB_IMIP_Tratamento/IMIP_SPIKE/"
]

results_dict = {}
counter = 1

for folder in folder_list:
    xlsx_path = folder / "results.xlsx"
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}")
        continue

    wb = load_workbook(filename=xlsx_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    i = 2  # Começa da terceira linha (índice 2)

    while i < len(rows):
        name_row = rows[i]
        score_row = rows[i + 1] if i + 1 < len(rows) else None

        if not name_row or not score_row:
            break

        # Ignora células vazias e junta o caminho
        components = [str(c).strip() for c in name_row if c and str(c).strip() != '']
        if components and components[0].startswith("File Name"):
            # Substitui o primeiro por "File Name: N" e junta o resto com /
            components[0] = f"File Name: {counter}"
            image_path = "/".join(components[1:])
            full_name = f"{components[0]} {image_path}"

            # Pega a 6ª coluna (índice 5) da linha de score
            score = score_row[5] if len(score_row) > 5 else None
            if score is not None:
                results_dict[full_name] = str(score).replace('.', ',')  # usa vírgula como separador decimal

            counter += 1
        i += 3  # pula para o próximo bloco

# Escreve no TXT
output_path = prefix / "extracted_scores.txt"
with open(output_path, "w", encoding="utf-8") as f:
    for key, val in results_dict.items():
        f.write(f"{key}\t{val}\n")

print(f"Output written to: {output_path}")

